import requests
import time
import json
import os
import random
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
import configparser
import re
import csv
import sys
import threading
import tkinter as tk
from tkinter import scrolledtext, messagebox, END, font as tkFont


# --- Global Constants ---
MASTER_FILE = "Etimad_Master_File.xlsx"
CONFIG_FILE = "config.ini"
TEMP_CSV_FILE = "new_projects.csv"

# --- Column Settings ---
ARABIC_HEADERS = [
    'إسم المشروع', 'الجهات', 'الربع السنوي', 'السنه',
    'مكان التنفيذ', 'طبيعة المشروع', 'وصف المشروع', 'الحالة',
    'مده التنفيذ المتوقعه (أيام)', 'مده التنفيذ المتوقعه (شهور)',
    'مده التنفيذ المتوقعه (سنين)', 'Encrypted_ID', 'تاريخ الإضافة'
]
UNIQUE_ID_COLUMN = 'Encrypted_ID'

# --- Scraper Settings ---
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36',
    'Referer': 'https://tenders.etimad.sa/Preplanning/PrePlaningForVisitor'
}
MAIN_PAGE_URL = "https://tenders.etimad.sa/Preplanning/PrePlaningForVisitor"
API_URL = "https://tenders.etimad.sa/Preplanning/PreplaningPagingForVisitorAsync"
PAGE_SIZE = 50


# --- Helper Functions ---

def clean_text(text):
    if text is None:
        return ""
    text = str(text)
    illegal_chars_re = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
    return illegal_chars_re.sub('', text)

def read_start_page():
    config = configparser.ConfigParser()
    start_page = 1
    if os.path.exists(CONFIG_FILE):
        try:
            config.read(CONFIG_FILE)
            start_page = int(config.get('Scraper', 'last_page', fallback=1))
            print(f" ✓ Configuration file found: {CONFIG_FILE}")
            print(f" ✓ Last successful page: {start_page}. Resuming from there.")
        except Exception as e:
            print(f" [!] Error reading {CONFIG_FILE}: {e}. Starting from page 1.")
            start_page = 1
    else:
        print(f" ✓ No configuration file found.")
        print(f" ✓ Starting full fetch from page 1 (this may take a while).")
    return start_page

def save_last_page(page_number):
    config = configparser.ConfigParser()
    config['Scraper'] = {'last_page': str(page_number)}
    try:
        with open(CONFIG_FILE, 'w') as f:
            config.write(f)
    except Exception as e:
        print(f"\n [!] Warning: Failed to save page {page_number} to {CONFIG_FILE}: {e}")

def append_csv_to_excel_and_delete(csv_file_path, excel_file_path):
    if not os.path.exists(csv_file_path):
        return 0
    print(f"\n 🔄 Merging temporary data from {csv_file_path} to {excel_file_path}...")
    if not os.path.exists(excel_file_path):
        print(f"   ✓ File {excel_file_path} does not exist. Creating with headers...")
        try:
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.append(ARABIC_HEADERS)
            ws_new.sheet_view.rightToLeft = True
            wb_new.save(excel_file_path)
            wb_new.close()
        except Exception as e:
            print(f" [!] Failed to create Excel file: {e}. Cannot continue.")
            return -1
    rows_to_add = []
    try:
        with open(csv_file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                rows_to_add.append(row)
    except Exception as e:
        print(f" [!] Critical error reading {csv_file_path}: {e}")
        return -1
    if not rows_to_add:
        print("   ✓ Temporary CSV file is empty. Deleting it.")
        try:
            os.remove(csv_file_path)
        except Exception as e:
            print(f" [!] Failed to delete empty CSV: {e}")
        return 0
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
        for row in rows_to_add:
            ws.append(row)
        wb.save(excel_file_path)
        wb.close()
        print(f"   ✓ Successfully merged {len(rows_to_add)} projects.")
        try:
            os.remove(csv_file_path)
            print(f"   ✓ Deleted temporary file {csv_file_path}.")
        except Exception as e:
            print(f" [!] Warning: Merge succeeded, but failed to delete {csv_file_path}: {e}")
        return len(rows_to_add)
    except PermissionError:
        print(f"\n{'!'*50}")
        print(f" [!] CRITICAL ERROR: File {excel_file_path} is currently open.")
        print(f" [!] Please close the Excel file and restart the tool.")
        print(f" [!] Data has NOT been lost. It's in {csv_file_path}.")
        print(f"{'!'*50}")
        messagebox.showerror("File Error",
                             f"File {excel_file_path} is currently open.\nPlease close it and restart.")
        return -1
    except Exception as e:
        print(f" [!] Critical error merging to Excel: {e}")
        return -1

def load_existing_project_ids_from_excel(filename):
    existing_ids = set()
    file_exists = os.path.exists(filename)
    total_existing = 0
    if file_exists:
        print(f"\n ✓ Found main Excel file: {filename}")
        print(f" ✓ Loading existing project IDs (to prevent duplicates)...")
        try:
            wb = load_workbook(filename, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            if not headers:
                wb.close()
                return set(), False, 0
            try:
                id_index = headers.index(UNIQUE_ID_COLUMN)
            except ValueError:
                print(f" [!] Error: Column '{UNIQUE_ID_COLUMN}' not found. Assuming new file.")
                wb.close()
                return set(), False, 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > id_index and row[id_index]:
                    existing_ids.add(str(row[id_index]))
                    total_existing += 1
            wb.close()
            print(f" ✓ Loaded {total_existing} existing project IDs.")
            return existing_ids, True, total_existing
        except Exception as e:
            print(f" [!] Error reading Excel file {filename}: {e}.")
            return set(), False, 0
    else:
        print(f"\n ✓ Main Excel file does not exist. Will create new file: {filename}")
        return set(), False, 0


# --- Main Scraper Function ---

def scrape_etimad_projects(on_complete_callback):
    try:
        start_time = time.time()
        print(f"{'='*50}")
        print(f"    Etimad Projects Scraper - Starting")
        print(f"    Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*50}")
        
        merge_result = append_csv_to_excel_and_delete(TEMP_CSV_FILE, MASTER_FILE)
        if merge_result == -1:
            print(" [!] Cancelled due to merge failure. Please review the log.")
            on_complete_callback()
            return
        
        start_page = read_start_page()
        existing_ids, file_exists, total_projects_before = load_existing_project_ids_from_excel(MASTER_FILE)
        
        session = requests.Session()
        session.headers.update(HEADERS)
        
        print("\n ✓ Visiting main page (to obtain valid cookies)...")
        try:
            session.get(MAIN_PAGE_URL, timeout=30).raise_for_status()
            print(f" ✓ Main page visit successful.")
        except requests.exceptions.RequestException as e:
            print(f" [!] Failed to visit main page: {e}. Exiting.")
            on_complete_callback()
            return
        
        try:
            csv_file = open(TEMP_CSV_FILE, 'a', encoding='utf-8-sig', newline='')
            csv_writer = csv.writer(csv_file)
        except Exception as e:
            print(f" [!] Critical error: Cannot open {TEMP_CSV_FILE} for writing: {e}")
            on_complete_callback()
            return
        
        print(f"\n--- Starting data fetch from page {start_page} ---")
        
        addition_date = datetime.now().strftime("%Y-%m-%d")
        current_page = start_page
        total_new_projects_added = 0
        start_page_for_summary = start_page
        last_page_processed = start_page - 1
        
        try:
            while True:
                timestamp = int(time.time() * 1000)
                params = {'pageSize': PAGE_SIZE, 'pageNumber': current_page, '_': timestamp}
                
                try:
                    response_api = session.get(API_URL, params=params, timeout=30)
                    response_api.raise_for_status()
                    data = response_api.json()
                    projects_list_json = data.get('data')
                    new_projects_on_this_page = 0
                    
                    if not projects_list_json:
                        print(f"\n  [✓✓] Reached end of data. No more pages after {current_page-1}.")
                        last_page_processed = current_page - 1
                        break
                    
                    for project_json in projects_list_json:
                        project_id = project_json.get('encyptedPrePlanningId', '')
                        if project_id and str(project_id) not in existing_ids:
                            new_projects_on_this_page += 1
                            total_new_projects_added += 1
                            row_data = [
                                clean_text(project_json.get('projectName', '')),
                                clean_text(project_json.get('agencyName', '')),
                                clean_text(project_json.get('yearQuarterName', '')),
                                clean_text(project_json.get('year', '')),
                                clean_text(project_json.get('insideKSAString', '')),
                                clean_text(project_json.get('projectNature', '')),
                                clean_text(project_json.get('projectDescription', '')),
                                clean_text(project_json.get('statusName', '')),
                                project_json.get('durationInDays', 0),
                                project_json.get('durationInMonths', 0),
                                project_json.get('durationInYears', 0),
                                clean_text(project_id),
                                clean_text(addition_date)
                            ]
                            csv_writer.writerow(row_data)
                            existing_ids.add(str(project_id))
                    
                    if new_projects_on_this_page > 0:
                        print(f"  ✓ Page {current_page}: Added {new_projects_on_this_page} new projects. (Total: {total_new_projects_added})")
                    else:
                        print(f"  ✓ Checked page {current_page}... (no new projects)")
                    
                    save_last_page(current_page)
                    last_page_processed = current_page
                    current_page += 1
                    
                    sleep_time = random.uniform(1.5, 4.0)
                    time.sleep(sleep_time)
                    
                except requests.exceptions.HTTPError as e:
                    if e.response.status_code == 429:
                        print(f"\n [!] Error 429 (rate limit) at page {current_page}. Waiting 60 seconds...")
                        time.sleep(60)
                    else:
                        print(f"\n [!] HTTP error {e.response.status_code} at page {current_page}. Waiting 10 seconds...")
                        time.sleep(10)
                except requests.exceptions.RequestException as e:
                    print(f"\n [!] Connection error ({e.__class__.__name__}) at page {current_page}. Waiting 15 seconds...")
                    time.sleep(15)
                except json.JSONDecodeError:
                    print(f"\n [!] Failed to parse JSON from page {current_page}. Retrying...")
                    time.sleep(5)
                    
        except KeyboardInterrupt:
            print(f"\n\n [!] Manual stop request received.")
            print(f" [!] Last saved page: {last_page_processed}")
        except Exception as e:
            print(f"\n [!] Critical error occurred: {e}")
            
    finally:
        print("\n ⏳ Closing temporary CSV file...")
        if 'csv_file' in locals() and not csv_file.closed:
            csv_file.close()
        
        print(" 🏁 Fetch complete. Merging collected data into main Excel file...")
        final_merge_count = append_csv_to_excel_and_delete(TEMP_CSV_FILE, MASTER_FILE)
        
        if final_merge_count == -1:
            print(" [!] Final merge failed. Data saved in new_projects.csv for next run.")
        
        print(f"\n\n{'='*50}")
        print(f"           📊 Update Summary 📊")
        print(f"{'='*50}")
        
        end_time = time.time()
        total_time_seconds = end_time - start_time
        total_time_str = time.strftime("%H:%M:%S", time.gmtime(total_time_seconds))
        print(f" ✓ Status: Operation completed.")
        print(f" ✓ Total time elapsed: {total_time_str}")
        print("-" * 50)
        
        total_pages_processed = max(0, (last_page_processed - start_page_for_summary + 1))
        print(f" ✓ Total pages checked: {total_pages_processed} pages")
        print(f"     (from page {start_page_for_summary} to page {last_page_processed})")
        print(f" ✓ Next run will start from page: {last_page_processed}")
        print("-" * 50)
        
        print(f" ✓ New projects added (this run): {total_new_projects_added}")
        print(f" ✓ Total projects before run: {total_projects_before}")
        
        initial_merge_count = merge_result if merge_result > 0 else 0
        total_projects_after = total_projects_before + total_new_projects_added + initial_merge_count
        print(f" ✓ New total projects in file: {total_projects_after}")
        print("-" * 50)
        
        print(f" ✓ Main Excel file updated: {os.path.abspath(MASTER_FILE)}")
        print(f" ✓ Configuration file updated: {os.path.abspath(CONFIG_FILE)}")
        print(f"{'='*50}")
        
        on_complete_callback()


# --- GUI Classes ---

class IORedirector(object):
    def __init__(self, text_widget):
        self.text_widget = text_widget
    
    def write(self, str):
        self.text_widget.after(0, self.insert_text, str)
    
    def insert_text(self, str):
        self.text_widget.config(state="normal")
        self.text_widget.insert(END, str)
        self.text_widget.see(END)
        self.text_widget.config(state="disabled")
    
    def flush(self):
        pass


class App:
    def __init__(self, root):
        self.root = root
        root.title("Etimad Projects")
        root.geometry("800x600")
        
        # Main frame
        self.main_frame = tk.Frame(root, padx=10, pady=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Fonts
        self.button_font = ("Arial", 12, "bold")
        self.label_font = ("Arial", 10)
        self.log_font = ("Courier New", 10)
        
        # Start button
        self.start_button = tk.Button(
            self.main_frame,
            text="Start",
            command=self.start_scraping_thread,
            font=self.button_font,
            bg="#4CAF50",
            fg="white",
            padx=10,
            pady=5
        )
        self.start_button.pack(pady=(0, 10), fill=tk.X)
        
        # Log label
        self.log_label = tk.Label(
            self.main_frame,
            text="Console Output:",
            font=self.label_font,
            anchor="w"
        )
        self.log_label.pack(fill=tk.X)
        
        # Log area - شاشة سوداء بنص أخضر Matrix style
        self.log_area = scrolledtext.ScrolledText(
            self.main_frame,
            height=10,
            wrap=tk.WORD,
            font=self.log_font,
            bg="#000000",
            fg="#FFD300",
            insertbackground="white"
        )
        self.log_area.config(state="disabled")
        self.log_area.pack(fill=tk.BOTH, expand=True)
        
        # Status bar
        self.status_label = tk.Label(
            root,
            text="Status: Ready",
            bd=1,
            relief=tk.SUNKEN,
            anchor="w",
            padx=10,
            font=self.label_font
        )
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Redirect print to GUI
        sys.stdout = IORedirector(self.log_area)
        sys.stderr = IORedirector(self.log_area)
        
        print("Welcome to Etimad Projects Scraper.")
        print("Click 'Start' button to begin fetching data.\n")
        
        root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.scraper_thread = None
    
    def start_scraping_thread(self):
        self.start_button.config(text="Running...", state="disabled", bg="#E0E0E0")
        self.status_label.config(text="Status: Running...")
        self.log_area.config(state="normal")
        self.log_area.delete('1.0', END)
        self.log_area.config(state="disabled")
        
        self.scraper_thread = threading.Thread(
            target=scrape_etimad_projects,
            args=(self.on_scraping_complete,),
            daemon=True
        )
        self.scraper_thread.start()
    
    def on_scraping_complete(self):
        self.root.after(0, self.update_gui_on_complete)
    
    def update_gui_on_complete(self):
        self.start_button.config(text="Start", state="normal", bg="#4CAF50")
        self.status_label.config(text="Status: Completed. Ready for next run.")
        messagebox.showinfo("Operation Complete", "Data fetching completed successfully!")
    
    def on_closing(self):
        if self.scraper_thread and self.scraper_thread.is_alive():
            if messagebox.askyesno("Confirm Exit", "The tool is still running.\nAre you sure you want to exit?"):
                self.root.destroy()
        else:
            self.root.destroy()


# --- Main Execution ---
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
